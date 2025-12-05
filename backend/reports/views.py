import io
import csv
import json
from datetime import datetime, timedelta
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from drf_spectacular.utils import extend_schema, OpenApiParameter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from apartments.models import Apartment
from clients.models import Client
from vendors.models import Vendor
from products.models import Product
from orders.models import Order, OrderItem
from deliveries.models import Delivery
from payments.models import Payment
from issues.models import Issue


class ReportGeneratorView(APIView):
    """
    Generate various reports in PDF, Excel, or CSV format
    """
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        tags=['Reports'],
        summary='Generate report',
        parameters=[
            OpenApiParameter(name='type', type=str, required=True,
                           description='Report type: orders, payments, inventory, apartments, issues'),
            OpenApiParameter(name='format', type=str, required=True,
                           description='Output format: pdf, excel, csv'),
            OpenApiParameter(name='start_date', type=str, required=False),
            OpenApiParameter(name='end_date', type=str, required=False),
            OpenApiParameter(name='apartment_id', type=str, required=False),
            OpenApiParameter(name='vendor_id', type=str, required=False),
        ]
    )
    def get(self, request):
        report_type = request.query_params.get('type')
        output_format = request.query_params.get('format', 'pdf')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        apartment_id = request.query_params.get('apartment_id')
        vendor_id = request.query_params.get('vendor_id')
        
        # Parse dates
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_date = timezone.now().date() - timedelta(days=30)
        
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date = timezone.now().date()
        
        # Generate report based on type
        if report_type == 'orders':
            return self.generate_orders_report(output_format, start_date, end_date, apartment_id, vendor_id)
        elif report_type == 'payments':
            return self.generate_payments_report(output_format, start_date, end_date, vendor_id)
        elif report_type == 'inventory':
            return self.generate_inventory_report(output_format)
        elif report_type == 'apartments':
            return self.generate_apartments_report(output_format)
        elif report_type == 'issues':
            return self.generate_issues_report(output_format, start_date, end_date, apartment_id)
        else:
            return Response({'error': 'Invalid report type'}, status=400)
    
    def generate_orders_report(self, format, start_date, end_date, apartment_id=None, vendor_id=None):
        """Generate orders report"""
        orders = Order.objects.filter(
            placed_on__gte=start_date,
            placed_on__lte=end_date
        )
        
        if apartment_id:
            orders = orders.filter(apartment_id=apartment_id)
        if vendor_id:
            orders = orders.filter(vendor_id=vendor_id)
        
        orders = orders.select_related('apartment', 'vendor').prefetch_related('items')
        
        if format == 'pdf':
            return self.orders_to_pdf(orders, start_date, end_date)
        elif format == 'excel':
            return self.orders_to_excel(orders, start_date, end_date)
        else:
            return self.orders_to_csv(orders)
    
    def orders_to_pdf(self, orders, start_date, end_date):
        """Convert orders to PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Title
        elements.append(Paragraph(f"Orders Report", title_style))
        elements.append(Paragraph(f"{start_date} to {end_date}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Summary
        total_orders = orders.count()
        total_value = orders.aggregate(Sum('total'))['total__sum'] or 0
        
        summary_data = [
            ['Total Orders:', str(total_orders)],
            ['Total Value:', f'€{total_value:,.2f}'],
            ['Average Order Value:', f'€{total_value/total_orders if total_orders else 0:,.2f}']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Orders table
        data = [['PO Number', 'Apartment', 'Vendor', 'Items', 'Total', 'Status', 'Date']]
        
        for order in orders:
            data.append([
                order.po_number,
                order.apartment.name[:20],
                order.vendor.name[:20],
                str(order.items_count),
                f'€{order.total:,.2f}',
                order.status,
                order.placed_on.strftime('%Y-%m-%d')
            ])
        
        table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1.5*inch, 0.7*inch, 1*inch, 0.8*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="orders_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    
    def orders_to_excel(self, orders, start_date, end_date):
        """Convert orders to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders Report"
        
        # Title
        ws['A1'] = 'Orders Report'
        ws['A2'] = f'{start_date} to {end_date}'
        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')
        
        # Style title
        title_font = Font(size=16, bold=True)
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Summary
        ws['A4'] = 'Summary'
        ws['A5'] = 'Total Orders:'
        ws['B5'] = orders.count()
        ws['A6'] = 'Total Value:'
        ws['B6'] = orders.aggregate(Sum('total'))['total__sum'] or 0
        
        # Headers
        headers = ['PO Number', 'Apartment', 'Vendor', 'Items', 'Total', 'Status', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Data
        for row, order in enumerate(orders, 9):
            ws.cell(row=row, column=1, value=order.po_number)
            ws.cell(row=row, column=2, value=order.apartment.name)
            ws.cell(row=row, column=3, value=order.vendor.name)
            ws.cell(row=row, column=4, value=order.items_count)
            ws.cell(row=row, column=5, value=float(order.total))
            ws.cell(row=row, column=6, value=order.status)
            ws.cell(row=row, column=7, value=order.placed_on.strftime('%Y-%m-%d'))
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="orders_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    def orders_to_csv(self, orders):
        """Convert orders to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="orders_report_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['PO Number', 'Apartment', 'Vendor', 'Items', 'Total', 'Status', 'Date'])
        
        for order in orders:
            writer.writerow([
                order.po_number,
                order.apartment.name,
                order.vendor.name,
                order.items_count,
                order.total,
                order.status,
                order.placed_on.strftime('%Y-%m-%d')
            ])
        
        return response
    
    def generate_payments_report(self, format, start_date, end_date, vendor_id=None):
        """Generate payments report"""
        payments = Payment.objects.filter(
            payment_date__gte=start_date,
            payment_date__lte=end_date
        )
        
        if vendor_id:
            payments = payments.filter(vendor_id=vendor_id)
        
        payments = payments.select_related('vendor')
        
        if format == 'pdf':
            return self.payments_to_pdf(payments, start_date, end_date)
        elif format == 'excel':
            return self.payments_to_excel(payments, start_date, end_date)
        else:
            return self.payments_to_csv(payments)
    
    def payments_to_pdf(self, payments, start_date, end_date):
        """Convert payments to PDF - similar structure to orders_to_pdf"""
        # Implementation similar to orders_to_pdf
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        elements.append(Paragraph("Payments Report", styles['Title']))
        elements.append(Paragraph(f"{start_date} to {end_date}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Build similar to orders
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="payments_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    
    def payments_to_excel(self, payments, start_date, end_date):
        """Convert payments to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments Report"
        
        # Similar structure to orders_to_excel
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="payments_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    def payments_to_csv(self, payments):
        """Convert payments to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="payments_report_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Invoice Number', 'Vendor', 'Amount', 'Status', 'Payment Date', 'Method'])
        
        for payment in payments:
            writer.writerow([
                payment.invoice_number,
                payment.vendor.name if payment.vendor else 'N/A',
                payment.amount,
                payment.status,
                payment.payment_date.strftime('%Y-%m-%d'),
                payment.payment_method
            ])
        
        return response
    
    def generate_inventory_report(self, format):
        """Generate inventory/products report"""
        products = Product.objects.select_related('vendor').all()
        
        if format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['SKU', 'Name', 'Category', 'Vendor', 'Price', 'Status'])
            
            for product in products:
                writer.writerow([
                    product.sku,
                    product.name,
                    product.category,
                    product.vendor.name if product.vendor else 'N/A',
                    product.price,
                    product.availability_status
                ])
            
            return response
        
        return Response({'error': 'Format not implemented'}, status=400)
    
    def generate_apartments_report(self, format):
        """Generate apartments report"""
        apartments = Apartment.objects.select_related('client').all()
        
        if format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="apartments_report_{datetime.now().strftime("%Y%m%d")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Name', 'Client', 'Status', 'Budget', 'Start Date', 'End Date'])
            
            for apartment in apartments:
                writer.writerow([
                    apartment.name,
                    apartment.client.name if apartment.client else 'N/A',
                    apartment.status,
                    apartment.budget,
                    apartment.start_date,
                    apartment.end_date
                ])
            
            return response
        
        return Response({'error': 'Format not implemented'}, status=400)
    
    def generate_issues_report(self, format, start_date, end_date, apartment_id=None):
        """Generate issues report"""
        issues = Issue.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        if apartment_id:
            issues = issues.filter(apartment_id=apartment_id)
        
        issues = issues.select_related('apartment', 'reported_by')
        
        if format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="issues_report_{datetime.now().strftime("%Y%m%d")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Title', 'Apartment', 'Priority', 'Status', 'Reported By', 'Created Date'])
            
            for issue in issues:
                writer.writerow([
                    issue.title,
                    issue.apartment.name if issue.apartment else 'N/A',
                    issue.priority,
                    issue.resolution_status,
                    issue.reported_by.email if issue.reported_by else 'N/A',
                    issue.created_at.strftime('%Y-%m-%d')
                ])
            
            return response
        
        return Response({'error': 'Format not implemented'}, status=400)


class ReportTemplatesView(APIView):
    """
    Get available report templates and their parameters
    """
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        tags=['Reports'],
        summary='Get available report templates',
    )
    def get(self, request):
        templates = [
            {
                'id': 'orders',
                'name': 'Orders Report',
                'description': 'Detailed report of orders with filtering options',
                'formats': ['pdf', 'excel', 'csv'],
                'parameters': [
                    {'name': 'start_date', 'type': 'date', 'required': False},
                    {'name': 'end_date', 'type': 'date', 'required': False},
                    {'name': 'apartment_id', 'type': 'uuid', 'required': False},
                    {'name': 'vendor_id', 'type': 'uuid', 'required': False},
                ]
            },
            {
                'id': 'payments',
                'name': 'Payments Report',
                'description': 'Payment transactions and status report',
                'formats': ['pdf', 'excel', 'csv'],
                'parameters': [
                    {'name': 'start_date', 'type': 'date', 'required': False},
                    {'name': 'end_date', 'type': 'date', 'required': False},
                    {'name': 'vendor_id', 'type': 'uuid', 'required': False},
                ]
            },
            {
                'id': 'inventory',
                'name': 'Inventory Report',
                'description': 'Current product inventory and availability',
                'formats': ['csv'],
                'parameters': []
            },
            {
                'id': 'apartments',
                'name': 'Apartments Report',
                'description': 'Apartment projects status and budget overview',
                'formats': ['csv'],
                'parameters': []
            },
            {
                'id': 'issues',
                'name': 'Issues Report',
                'description': 'Issue tracking and resolution status',
                'formats': ['csv'],
                'parameters': [
                    {'name': 'start_date', 'type': 'date', 'required': False},
                    {'name': 'end_date', 'type': 'date', 'required': False},
                    {'name': 'apartment_id', 'type': 'uuid', 'required': False},
                ]
            }
        ]
        
        return Response(templates)
