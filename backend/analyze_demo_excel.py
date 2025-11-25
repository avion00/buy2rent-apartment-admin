#!/usr/bin/env python
"""
Analyze the demo Excel file to identify image import issues
"""
import os
import pandas as pd
from openpyxl import load_workbook

def analyze_demo_excel():
    """Analyze the apartment-name-demo.xlsx file"""
    print("🔍 ANALYZING DEMO EXCEL FILE")
    print("=" * 35)
    
    excel_file = 'static/apartment-name-demo.xlsx'
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    print(f"✅ Found: {excel_file}")
    
    # Method 1: Pandas analysis
    print(f"\n📊 PANDAS ANALYSIS:")
    print("-" * 20)
    
    try:
        excel_data = pd.ExcelFile(excel_file)
        print(f"   📋 Sheets: {excel_data.sheet_names}")
        
        for sheet_name in excel_data.sheet_names:
            print(f"\n   📋 Sheet: {sheet_name}")
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            print(f"      • Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
            print(f"      • Original columns: {list(df.columns)}")
            
            # Show first few rows
            print(f"      • First 2 rows of data:")
            for i in range(min(2, len(df))):
                print(f"        Row {i+1}:")
                for col in df.columns:
                    value = df.iloc[i][col]
                    if pd.notna(value):
                        value_str = str(value)
                        if len(value_str) > 60:
                            value_str = value_str[:60] + "..."
                        print(f"          {col}: {value_str}")
            
            # Check for image-related columns
            image_columns = []
            for col in df.columns:
                col_lower = str(col).lower()
                if any(keyword in col_lower for keyword in ['image', 'photo', 'picture', 'url', 'link']):
                    image_columns.append(col)
            
            if image_columns:
                print(f"      • 🖼️  Image-related columns: {image_columns}")
                
                for img_col in image_columns:
                    print(f"        Column '{img_col}' content:")
                    for i in range(min(3, len(df))):
                        value = df.iloc[i][img_col]
                        if pd.notna(value):
                            print(f"          Row {i+1}: {value}")
                        else:
                            print(f"          Row {i+1}: [EMPTY]")
            else:
                print(f"      • ❌ No obvious image-related columns found")
                
                # Check all columns for URLs
                print(f"      • 🔍 Checking all columns for URLs:")
                for col in df.columns:
                    for i in range(min(2, len(df))):
                        value = df.iloc[i][col]
                        if pd.notna(value) and isinstance(value, str):
                            if 'http' in value.lower() or any(ext in value.lower() for ext in ['.jpg', '.png', '.jpeg']):
                                print(f"          Found URL in '{col}': {value}")
    
    except Exception as e:
        print(f"❌ Pandas analysis failed: {e}")
        import traceback
        traceback.print_exc()
    
    # Method 2: OpenpyXL analysis for embedded images
    print(f"\n🖼️  OPENPYXL ANALYSIS:")
    print("-" * 20)
    
    try:
        wb = load_workbook(excel_file, data_only=False)
        
        total_embedded = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n   📋 Sheet: {sheet_name}")
            
            # Check for embedded images
            if hasattr(ws, '_images') and ws._images:
                embedded_count = len(ws._images)
                total_embedded += embedded_count
                print(f"      • 🖼️  Embedded images: {embedded_count}")
                
                for i, img in enumerate(ws._images):
                    row_num = img.anchor._from.row + 1
                    col_num = img.anchor._from.col + 1
                    print(f"        Image {i+1}: Row {row_num}, Col {col_num}")
            else:
                print(f"      • No embedded images found")
        
        print(f"\n   📊 Total embedded images: {total_embedded}")
        
    except Exception as e:
        print(f"❌ OpenpyXL analysis failed: {e}")
    
    # Method 3: Column mapping test
    print(f"\n🗂️  COLUMN MAPPING TEST:")
    print("-" * 25)
    
    try:
        # Test with first sheet
        df = pd.read_excel(excel_file, sheet_name=excel_data.sheet_names[0])
        original_columns = list(df.columns)
        
        print(f"   Original columns: {original_columns}")
        
        # Normalize like import service does
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        normalized_columns = list(df.columns)
        
        print(f"   Normalized columns: {normalized_columns}")
        
        # Test our column mapping
        image_variations = ['product_image', 'product image', 'image', 'photo', 'picture', 'image_url', 'photo_url', 'picture_url']
        
        print(f"   Looking for: {image_variations}")
        
        found_matches = []
        for col in normalized_columns:
            if col in image_variations:
                found_matches.append(col)
        
        if found_matches:
            print(f"   ✅ Found matching columns: {found_matches}")
            
            # Check content
            for match in found_matches:
                print(f"   Content in '{match}':")
                for i in range(min(3, len(df))):
                    value = df.iloc[i][match]
                    print(f"      Row {i+1}: {value}")
        else:
            print(f"   ❌ NO MATCHING COLUMNS!")
            print(f"   🔧 Need to add column names to mapping")
            
            # Suggest which columns might be image columns
            potential = []
            for col in normalized_columns:
                if any(keyword in col for keyword in ['image', 'photo', 'picture', 'url', 'link']):
                    potential.append(col)
            
            if potential:
                print(f"   💡 Potential image columns: {potential}")
            else:
                print(f"   🤔 No obvious image columns found")
                print(f"   📋 All available columns: {normalized_columns}")
    
    except Exception as e:
        print(f"❌ Column mapping test failed: {e}")

def suggest_solution():
    """Suggest solution based on analysis"""
    print(f"\n💡 SOLUTION SUGGESTIONS:")
    print("=" * 25)
    
    print(f"Based on the analysis above:")
    print(f"1. If NO image columns found → Excel has no image data")
    print(f"2. If image columns found but not mapped → Update column mapping")
    print(f"3. If embedded images found → OpenpyXL extraction should work")
    print(f"4. If URLs found in unexpected columns → Add to mapping")
    
    print(f"\n🔧 QUICK FIXES:")
    print(f"• Add missing column names to import_service.py mapping")
    print(f"• Ensure Excel has image URLs in cells")
    print(f"• Check if images are embedded vs URL-based")

if __name__ == "__main__":
    print("🚀 DEMO EXCEL ANALYSIS\n")
    
    try:
        analyze_demo_excel()
        suggest_solution()
        
    except Exception as e:
        print(f"\n❌ Analysis failed: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print(f"\n✅ Analysis complete!")
