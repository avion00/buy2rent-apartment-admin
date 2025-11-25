#!/usr/bin/env python
"""
Analyze the exact structure of the Excel file to identify the issue
"""
import os
import pandas as pd
from openpyxl import load_workbook

def analyze_excel_structure():
    """Analyze Excel file structure in detail"""
    print("🔍 ANALYZING EXCEL FILE STRUCTURE")
    print("=" * 40)
    
    excel_file = 'sample_products_with_images.xlsx'
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    print(f"✅ Analyzing: {excel_file}")
    
    # Method 1: Pandas analysis
    print(f"\n📊 PANDAS ANALYSIS:")
    print("-" * 20)
    
    try:
        excel_data = pd.ExcelFile(excel_file)
        
        for sheet_name in excel_data.sheet_names:
            print(f"\n📋 Sheet: {sheet_name}")
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            print(f"   • Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
            print(f"   • Original columns: {list(df.columns)}")
            
            # Show first few rows
            print(f"   • First 3 rows:")
            for i in range(min(3, len(df))):
                print(f"     Row {i+1}:")
                for col in df.columns:
                    value = df.iloc[i][col]
                    if pd.notna(value):
                        value_str = str(value)
                        if len(value_str) > 50:
                            value_str = value_str[:50] + "..."
                        print(f"       {col}: {value_str}")
            
            # Check for image-related columns specifically
            image_columns = []
            for col in df.columns:
                col_lower = str(col).lower()
                if any(keyword in col_lower for keyword in ['image', 'photo', 'picture', 'url']):
                    image_columns.append(col)
            
            if image_columns:
                print(f"   • Image-related columns: {image_columns}")
                
                for img_col in image_columns:
                    print(f"     Column '{img_col}':")
                    for i in range(min(3, len(df))):
                        value = df.iloc[i][img_col]
                        print(f"       Row {i+1}: {value} (type: {type(value)})")
            else:
                print(f"   • No image-related columns found")
                
    except Exception as e:
        print(f"❌ Pandas analysis failed: {e}")
    
    # Method 2: OpenpyXL analysis
    print(f"\n🖼️  OPENPYXL ANALYSIS:")
    print("-" * 20)
    
    try:
        wb = load_workbook(excel_file, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n📋 Sheet: {sheet_name}")
            print(f"   • Max row: {ws.max_row}")
            print(f"   • Max column: {ws.max_column}")
            
            # Read headers
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(1, col).value
                if header:
                    headers.append(str(header))
            
            print(f"   • Headers: {headers}")
            
            # Check for embedded images
            if hasattr(ws, '_images') and ws._images:
                print(f"   • Embedded images: {len(ws._images)}")
                for i, img in enumerate(ws._images):
                    row_num = img.anchor._from.row + 1
                    col_num = img.anchor._from.col + 1
                    print(f"     Image {i+1}: Row {row_num}, Col {col_num}")
            else:
                print(f"   • No embedded images found")
            
            # Check cell values for image URLs
            print(f"   • Cell values (first 3 rows):")
            for row in range(1, min(4, ws.max_row + 1)):
                print(f"     Row {row}:")
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row, col).value
                    if cell_value:
                        cell_str = str(cell_value)
                        if len(cell_str) > 50:
                            cell_str = cell_str[:50] + "..."
                        header = headers[col-1] if col-1 < len(headers) else f"Col{col}"
                        print(f"       {header}: {cell_str}")
                        
                        # Check if it looks like an image URL
                        if isinstance(cell_value, str) and any(keyword in cell_value.lower() for keyword in ['http', 'image', '.jpg', '.png']):
                            print(f"         🖼️  ^ This looks like an image URL!")
                            
    except Exception as e:
        print(f"❌ OpenpyXL analysis failed: {e}")
    
    # Method 3: Column mapping test
    print(f"\n🗂️  COLUMN MAPPING TEST:")
    print("-" * 25)
    
    try:
        df = pd.read_excel(excel_file, sheet_name=excel_data.sheet_names[0])
        original_columns = list(df.columns)
        
        # Normalize columns like the import service does
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        normalized_columns = list(df.columns)
        
        print(f"   • Original → Normalized mapping:")
        for orig, norm in zip(original_columns, normalized_columns):
            print(f"     '{orig}' → '{norm}'")
        
        # Test the actual column mapping from import service
        column_mapping = {
            'product_image': ['product_image', 'product image', 'image', 'photo', 'picture', 'image_url', 'photo_url', 'picture_url'],
        }
        
        print(f"\n   • Testing image column mapping:")
        print(f"     Looking for: {column_mapping['product_image']}")
        print(f"     Available columns: {normalized_columns}")
        
        mapped_image_columns = []
        for col in normalized_columns:
            if col in column_mapping['product_image']:
                mapped_image_columns.append(col)
        
        if mapped_image_columns:
            print(f"     ✅ Mapped image columns: {mapped_image_columns}")
            
            # Test extraction from first row
            first_row = df.iloc[0]
            for img_col in mapped_image_columns:
                value = first_row[img_col]
                print(f"       {img_col}: '{value}' (pd.notna: {pd.notna(value)})")
        else:
            print(f"     ❌ NO IMAGE COLUMNS MAPPED!")
            print(f"     This is the problem - no columns match the mapping!")
            
    except Exception as e:
        print(f"❌ Column mapping test failed: {e}")

if __name__ == "__main__":
    analyze_excel_structure()
