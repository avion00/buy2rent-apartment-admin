import pandas as pd
import logging
import uuid
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import datetime
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from .models import Product
from .category_models import ProductCategory, ImportSession
from apartments.models import Apartment
from vendors.models import Vendor
import logging

logger = logging.getLogger(__name__)


class ProductImportService:
    """
    Service to handle Excel/CSV file imports for products
    """
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls', '.csv']
        self.image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']
    
    def validate_file(self, file):
        """Validate uploaded file"""
        errors = []
        
        # Check file extension
        file_ext = os.path.splitext(file.name)[1].lower()
        if file_ext not in self.supported_formats:
            errors.append(f"Unsupported file format. Supported: {', '.join(self.supported_formats)}")
        
        # Check file size (50MB limit)
        if file.size > 50 * 1024 * 1024:
            errors.append("File size exceeds 50MB limit")
        
        return errors
    
    def process_import(self, file, apartment_id, vendor_id=None, user=None):
        """
        Main method to process file import
        """
        try:
            # Validate file
            errors = self.validate_file(file)
            if errors:
                return {'success': False, 'errors': errors}
            
            # Get apartment
            apartment = Apartment.objects.get(id=apartment_id)
            
            # Get vendor if provided
            vendor = None
            if vendor_id:
                try:
                    vendor = Vendor.objects.get(id=vendor_id)
                    logger.info(f"✅ Vendor selected for import: {vendor.name} (ID: {vendor.id})")
                except Vendor.DoesNotExist:
                    return {'success': False, 'errors': ['Invalid vendor_id provided']}
            else:
                logger.warning("⚠️  No vendor_id provided for import")
            
            # Create import session and save the uploaded file permanently
            import_session = ImportSession.objects.create(
                apartment=apartment,
                file_name=file.name,
                file_size=file.size,
                file_type=os.path.splitext(file.name)[1].lower().replace('.', ''),
                uploaded_file=file,  # Save the file permanently
                status='processing'
            )
            
            logger.info(f"Saved uploaded file to: {import_session.uploaded_file.path}")
            
            # Use the saved file path for processing
            file_path = import_session.uploaded_file.path
            
            try:
                # Process file based on type
                if file.name.lower().endswith('.csv'):
                    result = self._process_csv(file_path, apartment, import_session, user, vendor)
                else:
                    result = self._process_excel_with_images(file_path, apartment, import_session, user, vendor)
                
                # Update import session
                import_session.status = 'completed' if result['success'] else 'failed'
                import_session.total_products = result.get('total_products', 0)
                import_session.successful_imports = result.get('successful_imports', 0)
                import_session.failed_imports = result.get('failed_imports', 0)
                import_session.completed_at = timezone.now()
                import_session.error_log = result.get('errors', [])
                import_session.save()
                
                return result
                
            except Exception as e:
                logger.error(f"Error during import processing: {str(e)}")
                import_session.status = 'failed'
                import_session.error_log = [str(e)]
                import_session.save()
                raise
                    
        except Exception as e:
            logger.error(f"Import error: {str(e)}")
            return {'success': False, 'errors': [str(e)]}
    
    def _save_temp_file(self, file):
        """Save uploaded file temporarily"""
        temp_dir = '/tmp'
        if not os.path.exists(temp_dir):
            temp_dir = '.'
        
        temp_filename = f"import_{uuid.uuid4()}_{file.name}"
        temp_path = os.path.join(temp_dir, temp_filename)
        
        with open(temp_path, 'wb') as temp_file:
            for chunk in file.chunks():
                temp_file.write(chunk)
        
        return temp_path
    
    def _process_csv(self, file_path, apartment, import_session, user, vendor=None):
        """Process CSV file"""
        try:
            df = pd.read_csv(file_path)
            
            # Create a single category for CSV
            category, created = ProductCategory.objects.get_or_create(
                apartment=apartment,
                sheet_name='CSV_Import',
                defaults={
                    'name': 'CSV Import',
                    'import_file_name': import_session.file_name,
                }
            )
            
            return self._process_dataframe(df, apartment, category, import_session, user, 'CSV_Import', vendor)
            
        except Exception as e:
            return {'success': False, 'errors': [f"CSV processing error: {str(e)}"]}
    
    def _process_excel(self, file_path, apartment, import_session, user):
        """Process Excel file with multiple sheets"""
        try:
            excel_file = pd.ExcelFile(file_path)
            
            total_products = 0
            successful_imports = 0
            failed_imports = 0
            all_errors = []
            
            import_session.total_sheets = len(excel_file.sheet_names)
            import_session.save()
            
            # Process each sheet
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Skip empty sheets
                    if df.empty or len(df) == 0:
                        logger.info(f"Skipping empty sheet: {sheet_name}")
                        continue
                    
                    # Create category for this sheet
                    category, created = ProductCategory.objects.get_or_create(
                        apartment=apartment,
                        sheet_name=sheet_name,
                        defaults={
                            'name': sheet_name.replace('_', ' ').title(),
                            'import_file_name': import_session.file_name,
                        }
                    )
                    
                    # Process dataframe
                    result = self._process_dataframe(df, apartment, category, import_session, user, sheet_name)
                    
                    total_products += result.get('total_products', 0)
                    successful_imports += result.get('successful_imports', 0)
                    failed_imports += result.get('failed_imports', 0)
                    all_errors.extend(result.get('errors', []))
                    
                except Exception as e:
                    error_msg = f"Sheet '{sheet_name}' error: {str(e)}"
                    all_errors.append(error_msg)
                    logger.error(error_msg)
            
            return {
                'success': len(all_errors) == 0,
                'total_products': total_products,
                'successful_imports': successful_imports,
                'failed_imports': failed_imports,
                'errors': all_errors,
                'sheets_processed': len(excel_file.sheet_names)
            }
            
        except Exception as e:
            return {'success': False, 'errors': [f"Excel processing error: {str(e)}"]}
    
    def _process_dataframe(self, df, apartment, category, import_session, user, sheet_name, vendor=None):
        """Process a pandas DataFrame and update products with vendor"""
        successful_imports = 0
        failed_imports = 0
        errors = []
        
        # Clean column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Map common column variations - Updated with all Excel columns
        column_mapping = {
            'sn': ['s.n', 'sn', 'serial_number', 'number', 'no'],
            'room': ['room', 'location', 'area'],
            'product_name': ['product_name', 'product', 'name', 'item', 'item_name', 'product name'],
            'product_image': ['product_image', 'product image', 'image', 'photo', 'picture', 'image_url', 'photo_url', 'picture_url'],
            'description': ['description', 'desc', 'details', 'product_description'],
            'sku': ['sku', 'product_code', 'item_code', 'code'],
            'quantity': ['quantity', 'qty', 'amount', 'count'],
            'cost': ['cost', 'price', 'unit_price'],
            'total_cost': ['total_cost', 'total cost', 'total_price', 'total price'],
            'link': ['link', 'url', 'vendor_link', 'product_link'],
            'size': ['size', 'dimensions', 'measurements'],
            'nm': ['nm', 'square_meter', 'sqm'],
            'plusz_nm': ['plusz_nm', 'plusz nm', 'plus_nm', 'plus nm', 'extra_nm'],
            'price_per_nm': ['price/nm', 'price_per_nm', 'price per nm'],
            'price_per_package': ['price/package', 'price_per_package', 'package_price'],
            'nm_per_package': ['nm/package', 'nm_per_package', 'nm per package'],
            'all_package': ['all_package', 'all package', 'total_packages'],
            'package_need_to_order': ['package_need_to_order', 'package need to order', 'packages_to_order'],
            'all_price': ['all_price', 'all price', 'total_amount', 'final_price'],
            'brand': ['brand', 'manufacturer', 'make'],
            'model': ['model', 'model_number', 'part_number'],
            'color': ['color', 'colour'],
            'material': ['material', 'fabric', 'composition'],
            'weight': ['weight'],
            'vendor_link': ['vendor_link', 'supplier_link'],
        }
        
        # Normalize columns
        normalized_columns = {}
        for standard_name, variations in column_mapping.items():
            for col in df.columns:
                if col in variations:
                    normalized_columns[col] = standard_name
                    break
        
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Check if row has meaningful data before processing
                    if not self._is_row_meaningful(row):
                        continue  # Skip empty/meaningless rows
                    
                    # Extract product data
                    product_data = self._extract_product_data(row, normalized_columns)
                    
                    # Create product with safe JSON data
                    import_data = {}
                    for key, value in row.to_dict().items():
                        if pd.notna(value):
                            import_data[str(key)] = str(value)
                        else:
                            import_data[str(key)] = None
                    
                    # Try to find existing product by SKU or product name
                    existing_product = None
                    if product_data.get('sku'):
                        existing_product = Product.objects.filter(
                            apartment=apartment,
                            sku=product_data['sku']
                        ).first()
                    
                    if not existing_product and product_data.get('product'):
                        existing_product = Product.objects.filter(
                            apartment=apartment,
                            product=product_data['product']
                        ).first()
                    
                    if existing_product:
                        # Update existing product
                        for key, value in product_data.items():
                            setattr(existing_product, key, value)
                        
                        # Update vendor - ALWAYS assign vendor if provided
                        if vendor:
                            existing_product.vendor = vendor
                            logger.info(f"✅ Assigned vendor '{vendor.name}' to existing product '{existing_product.product}'")
                        
                        existing_product.import_session = import_session
                        existing_product.import_row_number = index + 2
                        existing_product.import_data = import_data
                        existing_product.save()
                        
                        product = existing_product
                        logger.info(f"Updated existing product: {product.product} (SKU: {product.sku}) with vendor: {product.vendor.name if product.vendor else 'None'}")
                    else:
                        # Create new product
                        product = Product.objects.create(
                            apartment=apartment,
                            category=category,
                            vendor=vendor,
                            import_session=import_session,
                            import_row_number=index + 2,
                            import_data=import_data,
                            created_by=user.username if user else 'system',
                            **product_data
                        )
                        if vendor:
                            logger.info(f"✅ Created new product: {product.product} (SKU: {product.sku}) with vendor: {vendor.name}")
                        else:
                            logger.warning(f"⚠️  Created new product: {product.product} (SKU: {product.sku}) WITHOUT vendor")
                    
                    # Handle image if provided
                    if product_data.get('image_url'):
                        self._process_product_image(product, product_data['image_url'])
                    
                    successful_imports += 1
                    
                except Exception as e:
                    error_msg = f"Row {index + 2} in sheet '{sheet_name}': {str(e)}"
                    errors.append(error_msg)
                    failed_imports += 1
                    logger.error(error_msg)
        
        return {
            'success': len(errors) == 0,
            'total_products': len(df),
            'successful_imports': successful_imports,
            'failed_imports': failed_imports,
            'errors': errors
        }
    
    def _extract_product_data(self, row, column_mapping):
        """Extract product data from a row - Updated to handle all Excel columns"""
        data = {}
        
        # Required fields with defaults
        data['product'] = self._get_value(row, column_mapping, 'product_name', 'Unnamed Product')
        data['description'] = self._get_value(row, column_mapping, 'description', '')
        data['sku'] = self._get_value(row, column_mapping, 'sku', '')
        
        # Excel specific fields
        data['sn'] = self._get_value(row, column_mapping, 'sn', '')
        data['room'] = self._get_value(row, column_mapping, 'room', '')
        data['cost'] = self._get_value(row, column_mapping, 'cost', '')
        data['total_cost'] = self._get_value(row, column_mapping, 'total_cost', '')
        data['link'] = self._get_value(row, column_mapping, 'link', '')
        data['size'] = self._get_value(row, column_mapping, 'size', '')
        data['nm'] = self._get_value(row, column_mapping, 'nm', '')
        data['plusz_nm'] = self._get_value(row, column_mapping, 'plusz_nm', '')
        data['price_per_nm'] = self._get_value(row, column_mapping, 'price_per_nm', '')
        data['price_per_package'] = self._get_value(row, column_mapping, 'price_per_package', '')
        data['nm_per_package'] = self._get_value(row, column_mapping, 'nm_per_package', '')
        data['all_package'] = self._get_value(row, column_mapping, 'all_package', '')
        data['package_need_to_order'] = self._get_value(row, column_mapping, 'package_need_to_order', '')
        data['all_price'] = self._get_value(row, column_mapping, 'all_price', '')
        
        # Numeric fields - try to extract from cost field if available
        try:
            # Try to get price from cost field first, then fallback to unit_price
            cost_value = self._get_value(row, column_mapping, 'cost', 0)
            if cost_value and str(cost_value).strip():
                # Extract numeric value from cost string (e.g., "5000 Ft" -> 5000)
                import re
                price_match = re.search(r'[\d,]+', str(cost_value).replace(' ', ''))
                if price_match:
                    price_str = price_match.group().replace(',', '')
                    data['unit_price'] = float(price_str)
                else:
                    data['unit_price'] = 0
            else:
                data['unit_price'] = 0
        except (ValueError, TypeError):
            data['unit_price'] = 0
        
        try:
            qty = self._get_value(row, column_mapping, 'quantity', 1)
            data['qty'] = int(float(qty)) if qty and str(qty).strip() else 1
        except (ValueError, TypeError):
            data['qty'] = 1
        
        # Optional fields
        data['brand'] = self._get_value(row, column_mapping, 'brand', '')
        data['model_number'] = self._get_value(row, column_mapping, 'model', '')
        data['color'] = self._get_value(row, column_mapping, 'color', '')
        data['material'] = self._get_value(row, column_mapping, 'material', '')
        data['dimensions'] = self._get_value(row, column_mapping, 'size', '')  # Use size as dimensions
        data['weight'] = self._get_value(row, column_mapping, 'weight', '')
        # Handle image URLs - all image columns map to product_image only
        image_url = self._get_value(row, column_mapping, 'product_image', '')
        data['product_image'] = image_url  # Store in product_image field only
        data['vendor_link'] = self._get_value(row, column_mapping, 'vendor_link', '')
        
        return data
    
    def _get_value(self, row, column_mapping, field_name, default=''):
        """Get value from row using column mapping"""
        for col_name, mapped_name in column_mapping.items():
            if mapped_name == field_name and col_name in row.index:
                value = row[col_name]
                if pd.isna(value):
                    return default
                return str(value).strip()
        return default
    
    def _process_product_image(self, product, image_url):
        """Process product image - download and store locally or keep URL"""
        if not image_url or not image_url.strip():
            return
            
        try:
            # Check if it's a valid URL
            if image_url.startswith(('http://', 'https://')):
                # Option 1: Download and store locally (recommended)
                downloaded_url = self._download_and_store_image(product, image_url)
                if downloaded_url:
                    product.product_image = downloaded_url
                else:
                    # Fallback: store original URL if download fails
                    product.product_image = image_url
            else:
                # If it's not a URL, treat it as a filename or description
                product.product_image = image_url
                
            product.save(update_fields=['product_image'])
            
        except Exception as e:
            logger.error(f"Error processing image for product {product.id}: {str(e)}")
            # Store original URL as fallback
            product.product_image = image_url
            product.save(update_fields=['product_image'])
    
    def _download_and_store_image(self, product, image_url):
        """Download image from URL and store it locally"""
        try:
            import requests
            from django.core.files.base import ContentFile
            from django.core.files.storage import default_storage
            from urllib.parse import urlparse
            import mimetypes
            
            # Download the image
            response = requests.get(image_url, timeout=30, stream=True)
            response.raise_for_status()
            
            # Check if it's actually an image
            content_type = response.headers.get('content-type', '')
            if not content_type.startswith('image/'):
                logger.warning(f"URL {image_url} does not return an image (content-type: {content_type})")
                return None
            
            # Get file extension from URL or content type
            parsed_url = urlparse(image_url)
            file_extension = os.path.splitext(parsed_url.path)[1]
            if not file_extension:
                # Guess extension from content type
                file_extension = mimetypes.guess_extension(content_type) or '.jpg'
            
            # Generate unique filename
            filename = f"products/{product.apartment.id}/{product.id}_{uuid.uuid4().hex[:8]}{file_extension}"
            
            # Save the image
            image_content = ContentFile(response.content)
            saved_path = default_storage.save(filename, image_content)
            
            # Return the URL to access the saved image
            return default_storage.url(saved_path)
            
        except requests.RequestException as e:
            logger.error(f"Failed to download image from {image_url}: {str(e)}")
            return None
        except Exception as e:
            logger.error(f"Error storing image from {image_url}: {str(e)}")
            return None
    
    def _is_row_meaningful(self, row):
        """
        Check if a row has meaningful data worth importing.
        A row is considered meaningful if it has at least one of the key product fields.
        """
        # Key fields that indicate a meaningful product
        key_fields = [
            'product_name', 'product name', 'name',  # Product name variations
            'description', 'desc',                   # Description
            'cost', 'price', 'unit_price',          # Price information
            'quantity', 'qty',                       # Quantity
            'room', 'location',                      # Location
            'link', 'url', 'vendor_link',           # Links
            'brand', 'model',                        # Brand/model
            'size', 'dimensions',                    # Physical properties
            'material', 'color'                      # Material properties
        ]
        
        # Convert row to dict and check for meaningful values
        row_dict = row.to_dict()
        
        # Check if any key field has a non-empty, meaningful value
        for key, value in row_dict.items():
            if pd.notna(value) and str(value).strip():
                # Convert column name to lowercase for comparison
                col_name = str(key).lower().strip().replace(' ', '_')
                
                # Check if this column matches any key field
                for key_field in key_fields:
                    if key_field.replace(' ', '_') in col_name or col_name in key_field.replace(' ', '_'):
                        # Found a meaningful value in a key field
                        meaningful_value = str(value).strip()
                        
                        # Skip very short or generic values
                        if len(meaningful_value) >= 2 and meaningful_value.lower() not in ['n/a', 'na', '-', '0', '0.0']:
                            return True
        
        # If no meaningful data found in key fields, check if we have at least 3 non-empty fields
        non_empty_count = 0
        for key, value in row_dict.items():
            if pd.notna(value) and str(value).strip() and str(value).strip().lower() not in ['n/a', 'na', '-']:
                non_empty_count += 1
        
        # Require at least 3 non-empty fields for a row to be meaningful
        return non_empty_count >= 3
    
    def _extract_excel_images_with_openpyxl(self, file_path, apartment):
        """
        Extract images from Excel file using openpyxl to properly handle embedded images
        Returns a dictionary mapping row numbers to image file paths
        """
        try:
            from django.conf import settings
            
            # Load workbook with openpyxl
            wb = load_workbook(file_path, data_only=False)
            row_image_map = {}  # {sheet_name: {row_number: image_path}}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_images = {}
                
                # Extract images with row mapping
                if hasattr(ws, '_images') and ws._images:
                    logger.info(f"Found {len(ws._images)} images in sheet '{sheet_name}'")
                    
                    for i, img in enumerate(ws._images):
                        try:
                            # Get the row number from image anchor
                            row_num = img.anchor._from.row + 1  # Convert to 1-based indexing
                            logger.info(f"Processing image {i+1} at row {row_num} in sheet '{sheet_name}'")
                            
                            # Get image data first
                            if not hasattr(img, '_data'):
                                logger.warning(f"Image {i+1} in sheet '{sheet_name}' has no _data attribute, skipping")
                                continue
                            
                            img_data = img._data()
                            if not img_data:
                                logger.warning(f"Image {i+1} in sheet '{sheet_name}' has empty data, skipping")
                                continue
                            
                            # Determine image extension
                            if hasattr(img, 'format') and img.format:
                                extension = f".{img.format.lower()}"
                                logger.info(f"Image format from attribute: {extension}")
                            else:
                                # Try to detect format from image data
                                if img_data.startswith(b'\x89PNG'):
                                    extension = '.png'
                                elif img_data.startswith(b'\xff\xd8'):
                                    extension = '.jpg'
                                elif img_data.startswith(b'GIF'):
                                    extension = '.gif'
                                else:
                                    extension = '.png'  # Default
                                logger.info(f"Image format detected from data: {extension}")
                            
                            # Create media folder structure
                            folder_path = os.path.join(
                                settings.MEDIA_ROOT, 
                                'apartment_products', 
                                str(apartment.id),
                                sheet_name.replace(' ', '_').lower()
                            )
                            os.makedirs(folder_path, exist_ok=True)
                            logger.info(f"Created/verified folder: {folder_path}")
                            
                            # Generate unique filename
                            img_name = f"row_{row_num}_img_{i+1}_{uuid.uuid4().hex[:8]}{extension}"
                            img_path = os.path.join(folder_path, img_name)
                            
                            # Save image data
                            with open(img_path, 'wb') as f:
                                f.write(img_data)
                            
                            logger.info(f"Saved image to: {img_path}")
                            
                            # Store relative path for database
                            relative_path = os.path.join(
                                'apartment_products',
                                str(apartment.id),
                                sheet_name.replace(' ', '_').lower(),
                                img_name
                            ).replace('\\', '/')
                            
                            sheet_images[row_num] = f"/media/{relative_path}"
                            logger.info(f"✅ Extracted image for row {row_num}: {relative_path}")
                            
                        except Exception as e:
                            logger.error(f"❌ Error processing image {i+1} in sheet '{sheet_name}': {str(e)}", exc_info=True)
                            continue
                
                if sheet_images:
                    row_image_map[sheet_name] = sheet_images
                    logger.info(f"Sheet '{sheet_name}': Mapped {len(sheet_images)} images to rows")
            
            return row_image_map
            
        except Exception as e:
            logger.error(f"Error extracting images with openpyxl: {str(e)}")
            return {}
    
    def _process_excel_with_images(self, file_path, apartment, import_session, user, vendor=None):
        """
        Enhanced Excel processing that extracts embedded images using openpyxl
        """
        try:
            # First, extract all images from the Excel file
            logger.info("Extracting embedded images from Excel file...")
            row_image_map = self._extract_excel_images_with_openpyxl(file_path, apartment)
            
            # Then process the Excel file normally with pandas
            excel_file = pd.ExcelFile(file_path)
            
            total_products = 0
            successful_imports = 0
            failed_imports = 0
            all_errors = []
            
            import_session.total_sheets = len(excel_file.sheet_names)
            import_session.save()
            
            # Process each sheet
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Skip empty sheets
                    if df.empty or len(df) == 0:
                        logger.info(f"Skipping empty sheet: {sheet_name}")
                        continue
                    
                    # Create category for this sheet
                    category, created = ProductCategory.objects.get_or_create(
                        apartment=apartment,
                        sheet_name=sheet_name,
                        defaults={
                            'name': sheet_name.replace('_', ' ').title(),
                            'import_file_name': import_session.file_name,
                        }
                    )
                    
                    # Get images for this sheet
                    sheet_images = row_image_map.get(sheet_name, {})
                    
                    # Process dataframe with image mapping
                    result = self._process_dataframe_with_images(
                        df, apartment, category, import_session, user, sheet_name, sheet_images, vendor
                    )
                    
                    total_products += result.get('total_products', 0)
                    successful_imports += result.get('successful_imports', 0)
                    failed_imports += result.get('failed_imports', 0)
                    all_errors.extend(result.get('errors', []))
                    
                except Exception as e:
                    error_msg = f"Sheet '{sheet_name}' error: {str(e)}"
                    all_errors.append(error_msg)
                    logger.error(error_msg)
            
            return {
                'success': len(all_errors) == 0,
                'total_products': total_products,
                'successful_imports': successful_imports,
                'failed_imports': failed_imports,
                'errors': all_errors,
                'sheets_processed': len(excel_file.sheet_names),
                'images_extracted': sum(len(images) for images in row_image_map.values())
            }
            
        except Exception as e:
            return {'success': False, 'errors': [f"Excel with images processing error: {str(e)}"]}
    
    def _process_dataframe_with_images(self, df, apartment, category, import_session, user, sheet_name, sheet_images, vendor=None):
        """Process dataframe and assign images based on row mapping, update products with vendor"""
        successful_imports = 0
        failed_imports = 0
        errors = []
        
        # Clean column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Use existing column mapping
        column_mapping = {
            'sn': ['s.n', 'sn', 'serial_number', 'number', 'no'],
            'room': ['room', 'location', 'area'],
            'product_name': ['product_name', 'product', 'name', 'item', 'item_name', 'product name'],
            'product_image': ['product_image', 'product image', 'image', 'photo', 'picture', 'image_url', 'photo_url', 'picture_url'],
            'description': ['description', 'desc', 'details', 'product_description'],
            'sku': ['sku', 'product_code', 'item_code', 'code'],
            'quantity': ['quantity', 'qty', 'amount', 'count'],
            'cost': ['cost', 'price', 'unit_price'],
            'total_cost': ['total_cost', 'total cost', 'total_price', 'total price'],
            'link': ['link', 'url', 'vendor_link', 'product_link'],
            'size': ['size', 'dimensions', 'measurements'],
            'nm': ['nm', 'square_meter', 'sqm'],
            'plusz_nm': ['plusz_nm', 'plusz nm', 'plus_nm', 'plus nm', 'extra_nm'],
            'price_per_nm': ['price/nm', 'price_per_nm', 'price per nm'],
            'price_per_package': ['price/package', 'price_per_package', 'package_price'],
            'nm_per_package': ['nm/package', 'nm_per_package', 'nm per package'],
            'all_package': ['all_package', 'all package', 'total_packages'],
            'package_need_to_order': ['package_need_to_order', 'package need to order', 'packages_to_order'],
            'all_price': ['all_price', 'all price', 'total_amount', 'final_price'],
            'brand': ['brand', 'manufacturer', 'make'],
            'model': ['model', 'model_number', 'part_number'],
            'color': ['color', 'colour'],
            'material': ['material', 'fabric', 'composition'],
            'weight': ['weight'],
            'vendor_link': ['vendor_link', 'supplier_link'],
        }
        
        # Normalize columns
        normalized_columns = {}
        for standard_name, variations in column_mapping.items():
            for col in df.columns:
                if col in variations:
                    normalized_columns[col] = standard_name
                    break
        
        # Process each row
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if not self._is_row_meaningful(row):
                    continue
                
                # Extract product data
                product_data = self._extract_product_data(row, normalized_columns)
                
                # Create import data for reference
                import_data = {}
                for key, value in row.to_dict().items():
                    if pd.notna(value):
                        import_data[str(key)] = str(value)
                    else:
                        import_data[str(key)] = None
                
                # Try to find existing product by SKU or product name
                with transaction.atomic():
                    existing_product = None
                    if product_data.get('sku'):
                        existing_product = Product.objects.filter(
                            apartment=apartment,
                            sku=product_data['sku']
                        ).first()
                    
                    if not existing_product and product_data.get('product'):
                        existing_product = Product.objects.filter(
                            apartment=apartment,
                            product=product_data['product']
                        ).first()
                    
                    if existing_product:
                        # Update existing product
                        for key, value in product_data.items():
                            setattr(existing_product, key, value)
                        
                        # Update vendor - ALWAYS assign vendor if provided
                        if vendor:
                            existing_product.vendor = vendor
                            logger.info(f"✅ Assigned vendor '{vendor.name}' to existing product '{existing_product.product}'")
                        
                        existing_product.import_session = import_session
                        existing_product.import_row_number = index + 2
                        existing_product.import_data = import_data
                        existing_product.save()
                        
                        product = existing_product
                        logger.info(f"Updated existing product: {product.product} (SKU: {product.sku}) with vendor: {product.vendor.name if product.vendor else 'None'}")
                    else:
                        # Create new product
                        product = Product.objects.create(
                            apartment=apartment,
                            category=category,
                            vendor=vendor,
                            import_session=import_session,
                            import_row_number=index + 2,
                            import_data=import_data,
                            created_by=user.username if user else 'system',
                            **product_data
                        )
                        if vendor:
                            logger.info(f"✅ Created new product: {product.product} (SKU: {product.sku}) with vendor: {vendor.name}")
                        else:
                            logger.warning(f"⚠️  Created new product: {product.product} (SKU: {product.sku}) WITHOUT vendor")
                    
                    # Handle images: both embedded (from openpyxl) and URL-based (from cells)
                    excel_row = index + 2  # +2 because Excel has header row and is 1-based
                    
                    logger.info(f"Checking for images for product '{product.product}' at Excel row {excel_row}")
                    logger.info(f"Available images in sheet: {list(sheet_images.keys()) if sheet_images else 'None'}")
                    
                    # First, check for embedded images
                    if excel_row in sheet_images:
                        image_path = sheet_images[excel_row]
                        product.product_image = image_path
                        product.save(update_fields=['product_image'])
                        logger.info(f"✅ Assigned embedded image to product '{product.product}' (row {excel_row}): {image_path}")
                    
                    # Second, check for URL-based images from cells (if no embedded image found)
                    elif product_data.get('product_image'):
                        self._process_product_image(product, product_data['product_image'])
                        logger.info(f"Processing URL-based image for product '{product.product}': {product_data['product_image']}")
                    else:
                        logger.warning(f"⚠️  No image found for product '{product.product}' at Excel row {excel_row}")
                    
                    successful_imports += 1
                    
            except Exception as e:
                error_msg = f"Row {index + 2} in sheet '{sheet_name}': {str(e)}"
                errors.append(error_msg)
                failed_imports += 1
                logger.error(error_msg)
        
        return {
            'success': len(errors) == 0,
            'total_products': len(df),
            'successful_imports': successful_imports,
            'failed_imports': failed_imports,
            'errors': errors
        }
