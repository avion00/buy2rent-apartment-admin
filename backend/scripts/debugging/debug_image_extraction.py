import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from openpyxl import load_workbook
from apartments.models import Apartment
import pandas as pd

# Get the apartment
apartment_id = "7b78052b-5345-41ad-9942-cc2e6f8baf10"
apartment = Apartment.objects.get(id=apartment_id)

# Path to the Excel file - you'll need to update this
excel_file_path = input("Enter the path to your Excel file: ")

if not os.path.exists(excel_file_path):
    print(f"❌ File not found: {excel_file_path}")
    exit(1)

print(f"\n📊 Analyzing Excel file: {excel_file_path}")
print(f"🏢 Apartment: {apartment.name} ({apartment.id})")
print("=" * 60)

# Load with openpyxl to check images
wb = load_workbook(excel_file_path, data_only=False)

for sheet_name in wb.sheetnames:
    print(f"\n📄 Sheet: {sheet_name}")
    print("-" * 60)
    
    ws = wb[sheet_name]
    
    # Check for images
    if hasattr(ws, '_images') and ws._images:
        print(f"   ✅ Found {len(ws._images)} images")
        
        for i, img in enumerate(ws._images):
            try:
                row_num = img.anchor._from.row + 1
                col_num = img.anchor._from.col + 1
                print(f"   📷 Image {i+1}:")
                print(f"      • Row: {row_num} (openpyxl)")
                print(f"      • Column: {col_num}")
                print(f"      • Anchor type: {type(img.anchor).__name__}")
                if hasattr(img, 'format'):
                    print(f"      • Format: {img.format}")
            except Exception as e:
                print(f"   ❌ Error reading image {i+1}: {e}")
    else:
        print(f"   ⚠️  No images found in this sheet")
    
    # Load with pandas to check data
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
    print(f"\n   📋 Data rows: {len(df)}")
    print(f"   📋 Columns: {list(df.columns)}")
    
    # Show row mapping
    print(f"\n   🔢 Row number mapping:")
    for index, row in df.iterrows():
        excel_row = index + 2  # +2 for header and 0-based index
        product_name = row.get('Product Name', row.get('product name', 'N/A'))
        print(f"      • DataFrame index {index} → Excel row {excel_row} → Product: {product_name}")

print("\n" + "=" * 60)
print("✅ Analysis complete!")
