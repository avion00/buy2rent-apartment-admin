#!/usr/bin/env python
"""
Create Excel file with embedded images for testing openpyxl extraction
"""
import os
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import tempfile

def create_excel_with_embedded_images():
    """Create Excel file with actual embedded images"""
    print("🎨 CREATING EXCEL WITH EMBEDDED IMAGES")
    print("=" * 40)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Living Room sheet
    ws_living = wb.create_sheet("Living Room")
    
    # Add headers
    headers = ["S.N", "Product Name", "Room", "Quantity", "Cost", "Description"]
    for col, header in enumerate(headers, 1):
        ws_living.cell(1, col, header)
    
    # Add product data
    products = [
        ["1", "Modern Sofa", "Living Room", "1", "5000 Ft", "Comfortable 3-seater sofa"],
        ["2", "Coffee Table", "Living Room", "1", "1500 Ft", "Glass top coffee table"],
        ["3", "TV Stand", "Living Room", "1", "2000 Ft", "Modern TV entertainment unit"],
    ]
    
    for row, product in enumerate(products, 2):
        for col, value in enumerate(product, 1):
            ws_living.cell(row, col, value)
    
    # Create simple colored rectangles as "images" since we can't download real images easily
    print("🖼️  Creating test images...")
    
    try:
        # Create simple colored PNG images in memory
        from PIL import Image, ImageDraw
        
        colors = [
            (255, 107, 107),  # Red for sofa
            (150, 206, 180),  # Green for coffee table  
            (69, 183, 209),   # Blue for TV stand
        ]
        
        for i, (row, color) in enumerate(zip(range(2, 5), colors), 1):
            # Create a simple colored rectangle
            img = Image.new('RGB', (200, 150), color)
            draw = ImageDraw.Draw(img)
            
            # Add some text
            product_name = products[i-1][1]
            draw.text((10, 10), product_name, fill=(255, 255, 255))
            draw.text((10, 30), f"Product {i}", fill=(255, 255, 255))
            
            # Save to BytesIO
            img_buffer = BytesIO()
            img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Create openpyxl Image object
            openpyxl_img = OpenpyxlImage(img_buffer)
            
            # Set size
            openpyxl_img.width = 100
            openpyxl_img.height = 75
            
            # Add to worksheet at specific cell (next to the product data)
            cell_position = f"G{row}"  # Column G, row 2, 3, 4
            ws_living.add_image(openpyxl_img, cell_position)
            
            print(f"   ✅ Added image for {product_name} at {cell_position}")
    
    except ImportError:
        print("⚠️  PIL not available, creating placeholder images...")
        
        # Create simple text-based "images" using openpyxl drawing
        for i, row in enumerate(range(2, 5), 1):
            # We'll just add a note that an image should be here
            ws_living.cell(row, 7, f"[IMAGE_{i}]")
    
    except Exception as e:
        print(f"❌ Error creating images: {e}")
        # Add placeholder text
        for i, row in enumerate(range(2, 5), 1):
            ws_living.cell(row, 7, f"[IMAGE_{i}]")
    
    # Create Bedroom sheet
    ws_bedroom = wb.create_sheet("Bedroom")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        ws_bedroom.cell(1, col, header)
    
    # Add bedroom products
    bedroom_products = [
        ["4", "Queen Bed", "Bedroom", "1", "8000 Ft", "Comfortable queen size bed"],
        ["5", "Nightstand", "Bedroom", "2", "1200 Ft", "Wooden bedside table"],
    ]
    
    for row, product in enumerate(bedroom_products, 2):
        for col, value in enumerate(product, 1):
            ws_bedroom.cell(row, col, value)
    
    # Add images to bedroom sheet too
    try:
        from PIL import Image, ImageDraw
        
        bedroom_colors = [
            (220, 160, 220),  # Purple for bed
            (240, 230, 140),  # Khaki for nightstand
        ]
        
        for i, (row, color) in enumerate(zip(range(2, 4), bedroom_colors), 1):
            # Create a simple colored rectangle
            img = Image.new('RGB', (200, 150), color)
            draw = ImageDraw.Draw(img)
            
            # Add some text
            product_name = bedroom_products[i-1][1]
            draw.text((10, 10), product_name, fill=(0, 0, 0))
            draw.text((10, 30), f"Bedroom Item {i}", fill=(0, 0, 0))
            
            # Save to BytesIO
            img_buffer = BytesIO()
            img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Create openpyxl Image object
            openpyxl_img = OpenpyxlImage(img_buffer)
            
            # Set size
            openpyxl_img.width = 100
            openpyxl_img.height = 75
            
            # Add to worksheet
            cell_position = f"G{row}"
            ws_bedroom.add_image(openpyxl_img, cell_position)
            
            print(f"   ✅ Added image for {product_name} at {cell_position}")
    
    except:
        # Add placeholder text
        for i, row in enumerate(range(2, 4), 1):
            ws_bedroom.cell(row, 7, f"[BEDROOM_IMAGE_{i}]")
    
    # Save workbook
    output_file = "sample_products_with_embedded_images.xlsx"
    wb.save(output_file)
    
    print(f"\n✅ Created Excel file: {output_file}")
    
    # Verify the file
    try:
        from openpyxl import load_workbook
        
        # Load and check
        wb_test = load_workbook(output_file, data_only=False)
        total_images = 0
        
        for sheet_name in wb_test.sheetnames:
            ws = wb_test[sheet_name]
            if hasattr(ws, '_images') and ws._images:
                sheet_images = len(ws._images)
                total_images += sheet_images
                print(f"   📋 {sheet_name}: {sheet_images} embedded images")
        
        print(f"   🖼️  Total embedded images: {total_images}")
        
        if total_images > 0:
            print(f"\n🎉 SUCCESS! Excel file created with {total_images} embedded images")
            print(f"   Ready for openpyxl extraction testing")
        else:
            print(f"\n⚠️  No embedded images detected in created file")
            print(f"   This might be due to PIL not being available")
        
        return output_file, total_images > 0
        
    except Exception as e:
        print(f"❌ Error verifying file: {e}")
        return output_file, False

if __name__ == "__main__":
    print("🚀 EXCEL WITH EMBEDDED IMAGES CREATOR\n")
    
    try:
        output_file, has_images = create_excel_with_embedded_images()
        
        if has_images:
            print(f"\n🎯 NEXT STEPS:")
            print(f"1. Run: python test_openpyxl_images.py")
            print(f"2. Test import with: {output_file}")
            print(f"3. Check if images are extracted and saved")
        else:
            print(f"\n💡 ALTERNATIVE:")
            print(f"1. Use the existing sample_products_with_images.xlsx (has URLs)")
            print(f"2. The original URL-based import should work")
            print(f"3. OpenpyXL extraction is for embedded images only")
        
    except Exception as e:
        print(f"\n❌ Creation failed: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print(f"\n✅ Creation complete!")
