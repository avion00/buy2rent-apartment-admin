import pandas as pd
import logging
import uuid
import os
from openpyxl import load_workbook
from datetime import datetime, date
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from .models import Order, OrderItem
from products.models import Product
from apartments.models import Apartment
from vendors.models import Vendor

logger = logging.getLogger(__name__)


class OrderImportService:
    """
    Service to handle Excel/CSV file imports for orders
    Creates an order and imports all products from the file as order items
    """
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls', '.csv']
        self.image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']
    
    def validate_file(self, file):
        """Validate uploaded file"""
        errors = []
        
        # Check file extension
        file_ext = os.path.splitext(file.name)[1].lower()
        if file_ext not in self.supported_formats:
            errors.append(f"Unsupported file format. Supported: {', '.join(self.supported_formats)}")
        
        # Check file size (50MB limit)
        if file.size > 50 * 1024 * 1024:
            errors.append("File size exceeds 50MB limit")
        
        return errors
    
    def process_import(self, file, order_data, user=None):
        """
        Main method to process file import and create order with items
        
        Args:
            file: Uploaded Excel/CSV file
            order_data: Dictionary containing order information
                - apartment_id: UUID of apartment
                - vendor_id: UUID of vendor
                - po_number: Purchase order number
                - status: Order status (default: 'draft')
                - confirmation_code: Optional confirmation code
                - tracking_number: Optional tracking number
                - expected_delivery: Optional expected delivery date
                - shipping_address: Optional shipping address
                - notes: Optional notes
            user: User object
        
        Returns:
            Dictionary with success status and results
        """
        try:
            # Validate file
            errors = self.validate_file(file)
            if errors:
                return {'success': False, 'errors': errors}
            
            # Get apartment
            apartment_id = order_data.get('apartment_id')
            if not apartment_id:
                return {'success': False, 'errors': ['apartment_id is required']}
            
            try:
                apartment = Apartment.objects.get(id=apartment_id)
            except Apartment.DoesNotExist:
                return {'success': False, 'errors': ['Invalid apartment_id provided']}
            
            # Get vendor
            vendor_id = order_data.get('vendor_id')
            if not vendor_id:
                return {'success': False, 'errors': ['vendor_id is required']}
            
            try:
                vendor = Vendor.objects.get(id=vendor_id)
            except Vendor.DoesNotExist:
                return {'success': False, 'errors': ['Invalid vendor_id provided']}
            
            # Save file temporarily
            temp_path = self._save_temp_file(file)
            
            try:
                # Parse the file to get products
                if file.name.lower().endswith('.csv'):
                    products_data = self._parse_csv(temp_path)
                else:
                    products_data = self._parse_excel(temp_path, apartment)
                
                if not products_data:
                    return {'success': False, 'errors': ['No valid products found in file']}
                
                # Create order and order items in a transaction
                result = self._create_order_with_items(
                    apartment=apartment,
                    vendor=vendor,
                    products_data=products_data,
                    order_data=order_data,
                    user=user
                )
                
                return result
                
            finally:
                # Clean up temp file
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                    
        except Exception as e:
            logger.error(f"Order import error: {str(e)}", exc_info=True)
            return {'success': False, 'errors': [str(e)]}
    
    def _save_temp_file(self, file):
        """Save uploaded file temporarily"""
        temp_dir = '/tmp'
        if not os.path.exists(temp_dir):
            temp_dir = '.'
        
        temp_filename = f"order_import_{uuid.uuid4()}_{file.name}"
        temp_path = os.path.join(temp_dir, temp_filename)
        
        with open(temp_path, 'wb') as temp_file:
            for chunk in file.chunks():
                temp_file.write(chunk)
        
        return temp_path
    
    def _parse_csv(self, file_path):
        """Parse CSV file and extract product data"""
        try:
            df = pd.read_csv(file_path)
            return self._extract_products_from_dataframe(df)
        except Exception as e:
            logger.error(f"CSV parsing error: {str(e)}")
            return []
    
    def _parse_excel(self, file_path, apartment):
        """Parse Excel file and extract product data from all sheets"""
        try:
            # Extract images first
            row_image_map = self._extract_excel_images_with_openpyxl(file_path, apartment)
            
            excel_file = pd.ExcelFile(file_path)
            all_products = []
            
            # Process each sheet
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Skip empty sheets
                    if df.empty or len(df) == 0:
                        logger.info(f"Skipping empty sheet: {sheet_name}")
                        continue
                    
                    # Get images for this sheet
                    sheet_images = row_image_map.get(sheet_name, {})
                    
                    # Extract products from this sheet
                    products = self._extract_products_from_dataframe(df, sheet_images)
                    all_products.extend(products)
                    
                except Exception as e:
                    logger.error(f"Error processing sheet '{sheet_name}': {str(e)}")
                    continue
            
            return all_products
            
        except Exception as e:
            logger.error(f"Excel parsing error: {str(e)}")
            return []
    
    def _extract_products_from_dataframe(self, df, sheet_images=None):
        """Extract product data from a pandas DataFrame"""
        products = []
        
        # Clean column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Map common column variations
        column_mapping = {
            'sn': ['s.n', 'sn', 'serial_number', 'number', 'no'],
            'room': ['room', 'location', 'area'],
            'product_name': ['product_name', 'product', 'name', 'item', 'item_name', 'product name'],
            'product_image': ['product_image', 'product image', 'image', 'photo', 'picture', 'image_url', 'photo_url', 'picture_url'],
            'description': ['description', 'desc', 'details', 'product_description'],
            'sku': ['sku', 'product_code', 'item_code', 'code'],
            'quantity': ['quantity', 'qty', 'amount', 'count'],
            'cost': ['cost', 'price', 'unit_price'],
            'total_cost': ['total_cost', 'total cost', 'total_price', 'total price'],
            'link': ['link', 'url', 'vendor_link', 'product_link'],
            'size': ['size', 'dimensions', 'measurements'],
            'brand': ['brand', 'manufacturer', 'make'],
            'model': ['model', 'model_number', 'part_number'],
            'color': ['color', 'colour'],
            'material': ['material', 'fabric', 'composition'],
            'weight': ['weight'],
        }
        
        # Normalize columns
        normalized_columns = {}
        for standard_name, variations in column_mapping.items():
            for col in df.columns:
                if col in variations:
                    normalized_columns[col] = standard_name
                    break
        
        # Process each row
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if not self._is_row_meaningful(row):
                    continue
                
                # Extract product data
                product_data = self._extract_product_data(row, normalized_columns)
                
                # Add image from Excel if available
                if sheet_images:
                    excel_row = index + 2  # +2 because Excel has header row and is 1-based
                    if excel_row in sheet_images:
                        product_data['product_image'] = sheet_images[excel_row]
                
                products.append(product_data)
                
            except Exception as e:
                logger.error(f"Error processing row {index + 2}: {str(e)}")
                continue
        
        return products
    
    def _extract_product_data(self, row, column_mapping):
        """Extract product data from a row"""
        data = {}
        
        # Required fields with defaults
        data['product_name'] = self._get_value(row, column_mapping, 'product_name', 'Unnamed Product')
        data['description'] = self._get_value(row, column_mapping, 'description', '')
        data['sku'] = self._get_value(row, column_mapping, 'sku', '')
        
        # Numeric fields
        try:
            cost_value = self._get_value(row, column_mapping, 'cost', 0)
            if cost_value and str(cost_value).strip():
                # Extract numeric value from cost string (e.g., "5000 Ft" -> 5000)
                import re
                price_match = re.search(r'[\d,.]+', str(cost_value).replace(' ', ''))
                if price_match:
                    price_str = price_match.group().replace(',', '')
                    data['unit_price'] = float(price_str)
                else:
                    data['unit_price'] = 0
            else:
                data['unit_price'] = 0
        except (ValueError, TypeError):
            data['unit_price'] = 0
        
        try:
            qty = self._get_value(row, column_mapping, 'quantity', 1)
            data['quantity'] = int(float(qty)) if qty and str(qty).strip() else 1
        except (ValueError, TypeError):
            data['quantity'] = 1
        
        # Optional fields
        data['brand'] = self._get_value(row, column_mapping, 'brand', '')
        data['model'] = self._get_value(row, column_mapping, 'model', '')
        data['color'] = self._get_value(row, column_mapping, 'color', '')
        data['material'] = self._get_value(row, column_mapping, 'material', '')
        data['size'] = self._get_value(row, column_mapping, 'size', '')
        data['weight'] = self._get_value(row, column_mapping, 'weight', '')
        data['product_image'] = self._get_value(row, column_mapping, 'product_image', '')
        data['link'] = self._get_value(row, column_mapping, 'link', '')
        
        # Build specifications JSON
        specifications = {}
        if data['brand']:
            specifications['brand'] = data['brand']
        if data['model']:
            specifications['model'] = data['model']
        if data['color']:
            specifications['color'] = data['color']
        if data['material']:
            specifications['material'] = data['material']
        if data['size']:
            specifications['size'] = data['size']
        if data['weight']:
            specifications['weight'] = data['weight']
        
        data['specifications'] = specifications
        
        return data
    
    def _get_value(self, row, column_mapping, field_name, default=''):
        """Get value from row using column mapping"""
        for col_name, mapped_name in column_mapping.items():
            if mapped_name == field_name and col_name in row.index:
                value = row[col_name]
                if pd.isna(value):
                    return default
                return str(value).strip()
        return default
    
    def _is_row_meaningful(self, row):
        """Check if a row has meaningful data worth importing"""
        key_fields = [
            'product_name', 'product name', 'name',
            'description', 'desc',
            'cost', 'price', 'unit_price',
            'quantity', 'qty',
        ]
        
        row_dict = row.to_dict()
        
        for key, value in row_dict.items():
            if pd.notna(value) and str(value).strip():
                col_name = str(key).lower().strip().replace(' ', '_')
                
                for key_field in key_fields:
                    if key_field.replace(' ', '_') in col_name or col_name in key_field.replace(' ', '_'):
                        meaningful_value = str(value).strip()
                        
                        if len(meaningful_value) >= 2 and meaningful_value.lower() not in ['n/a', 'na', '-', '0', '0.0']:
                            return True
        
        return False
    
    def _extract_excel_images_with_openpyxl(self, file_path, apartment):
        """Extract images from Excel file using openpyxl"""
        try:
            from django.conf import settings
            
            wb = load_workbook(file_path, data_only=False)
            row_image_map = {}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_images = {}
                
                if hasattr(ws, '_images') and ws._images:
                    logger.info(f"Found {len(ws._images)} images in sheet '{sheet_name}'")
                    
                    for i, img in enumerate(ws._images):
                        try:
                            row_num = img.anchor._from.row + 1
                            
                            if not hasattr(img, '_data'):
                                continue
                            
                            img_data = img._data()
                            if not img_data:
                                continue
                            
                            # Determine image extension
                            if hasattr(img, 'format') and img.format:
                                extension = f".{img.format.lower()}"
                            else:
                                if img_data.startswith(b'\x89PNG'):
                                    extension = '.png'
                                elif img_data.startswith(b'\xff\xd8'):
                                    extension = '.jpg'
                                elif img_data.startswith(b'GIF'):
                                    extension = '.gif'
                                else:
                                    extension = '.png'
                            
                            # Create media folder structure
                            folder_path = os.path.join(
                                settings.MEDIA_ROOT,
                                'order_products',
                                str(apartment.id),
                                sheet_name.replace(' ', '_').lower()
                            )
                            os.makedirs(folder_path, exist_ok=True)
                            
                            # Generate unique filename
                            img_name = f"row_{row_num}_img_{i+1}_{uuid.uuid4().hex[:8]}{extension}"
                            img_path = os.path.join(folder_path, img_name)
                            
                            # Save image data
                            with open(img_path, 'wb') as f:
                                f.write(img_data)
                            
                            # Store relative path for database
                            relative_path = os.path.join(
                                'order_products',
                                str(apartment.id),
                                sheet_name.replace(' ', '_').lower(),
                                img_name
                            ).replace('\\', '/')
                            
                            sheet_images[row_num] = f"/media/{relative_path}"
                            logger.info(f"✅ Extracted image for row {row_num}: {relative_path}")
                            
                        except Exception as e:
                            logger.error(f"❌ Error processing image {i+1} in sheet '{sheet_name}': {str(e)}")
                            continue
                
                if sheet_images:
                    row_image_map[sheet_name] = sheet_images
            
            return row_image_map
            
        except Exception as e:
            logger.error(f"Error extracting images with openpyxl: {str(e)}")
            return {}
    
    def _create_order_with_items(self, apartment, vendor, products_data, order_data, user):
        """Create order and order items in a transaction"""
        try:
            with transaction.atomic():
                # Calculate totals
                total_amount = sum(p['unit_price'] * p['quantity'] for p in products_data)
                items_count = len(products_data)
                
                # Parse dates
                placed_on = order_data.get('placed_on')
                if not placed_on:
                    placed_on = date.today()
                elif isinstance(placed_on, str):
                    placed_on = datetime.strptime(placed_on, '%Y-%m-%d').date()
                
                expected_delivery = order_data.get('expected_delivery')
                if expected_delivery and isinstance(expected_delivery, str):
                    expected_delivery = datetime.strptime(expected_delivery, '%Y-%m-%d').date()
                
                # Create order
                order = Order.objects.create(
                    po_number=order_data.get('po_number'),
                    apartment=apartment,
                    vendor=vendor,
                    items_count=items_count,
                    total=total_amount,
                    status=order_data.get('status', 'draft'),
                    confirmation_code=order_data.get('confirmation_code', ''),
                    placed_on=placed_on,
                    expected_delivery=expected_delivery,
                    tracking_number=order_data.get('tracking_number', ''),
                    shipping_address=order_data.get('shipping_address', ''),
                    notes=order_data.get('notes', ''),
                )
                
                logger.info(f"✅ Created order: {order.po_number} (ID: {order.id})")
                
                # Create order items and link to products
                successful_items = 0
                failed_items = 0
                errors = []
                
                for product_data in products_data:
                    try:
                        # Try to find existing product by SKU or name
                        product = None
                        if product_data.get('sku'):
                            product = Product.objects.filter(
                                apartment=apartment,
                                sku=product_data['sku']
                            ).first()
                        
                        if not product and product_data.get('product_name'):
                            product = Product.objects.filter(
                                apartment=apartment,
                                product=product_data['product_name']
                            ).first()
                        
                        # Create product if it doesn't exist
                        if not product:
                            product = Product.objects.create(
                                apartment=apartment,
                                vendor=vendor,
                                product=product_data['product_name'],
                                sku=product_data.get('sku', ''),
                                unit_price=product_data['unit_price'],
                                qty=product_data['quantity'],
                                description=product_data.get('description', ''),
                                product_image=product_data.get('product_image', ''),
                                status=['Ordered'],  # Mark as ordered since it's from an order import
                                availability='In Stock',
                            )
                            logger.info(f"✅ Created new product: {product.product} (SKU: {product.sku})")
                        
                        # Create order item
                        order_item = OrderItem.objects.create(
                            order=order,
                            product=product,  # Now always has a product reference
                            product_name=product_data['product_name'],
                            product_image_url=product_data.get('product_image', ''),
                            sku=product_data.get('sku', ''),
                            quantity=product_data['quantity'],
                            unit_price=product_data['unit_price'],
                            total_price=product_data['unit_price'] * product_data['quantity'],
                            description=product_data.get('description', ''),
                            specifications=product_data.get('specifications', {}),
                        )
                        
                        successful_items += 1
                        logger.info(f"✅ Created order item: {order_item.product_name} (Qty: {order_item.quantity})")
                        
                    except Exception as e:
                        error_msg = f"Failed to create order item for '{product_data.get('product_name', 'Unknown')}': {str(e)}"
                        errors.append(error_msg)
                        failed_items += 1
                        logger.error(error_msg)
                
                return {
                    'success': True,
                    'message': 'Order and items imported successfully',
                    'data': {
                        'order_id': str(order.id),
                        'po_number': order.po_number,
                        'total_items': items_count,
                        'successful_imports': successful_items,
                        'failed_imports': failed_items,
                        'total_amount': float(total_amount),
                        'errors': errors
                    }
                }
                
        except Exception as e:
            logger.error(f"Error creating order with items: {str(e)}", exc_info=True)
            return {
                'success': False,
                'errors': [f"Failed to create order: {str(e)}"]
            }
