#!/usr/bin/env python
"""
Test openpyxl image extraction from Excel files
"""
import os
import sys
import django

# Set up Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from openpyxl import load_workbook
from products.import_service import ProductImportService
from apartments.models import Apartment
from clients.models import Client

def test_openpyxl_extraction():
    """Test openpyxl image extraction"""
    print("🧪 TESTING OPENPYXL IMAGE EXTRACTION")
    print("=" * 40)
    
    # Check if sample Excel exists
    excel_file = 'sample_products_with_images.xlsx'
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False
    
    print(f"✅ Found Excel file: {excel_file}")
    
    # Test openpyxl workbook loading
    try:
        wb = load_workbook(excel_file, data_only=False)
        print(f"✅ Loaded workbook with {len(wb.sheetnames)} sheets: {wb.sheetnames}")
        
        total_images = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n📋 Sheet: {sheet_name}")
            
            # Check for images
            if hasattr(ws, '_images') and ws._images:
                images = ws._images
                print(f"   🖼️  Found {len(images)} embedded images")
                total_images += len(images)
                
                for i, img in enumerate(images):
                    try:
                        # Get anchor information
                        anchor = img.anchor
                        row_num = anchor._from.row + 1  # Convert to 1-based
                        col_num = anchor._from.col + 1
                        
                        print(f"      Image {i+1}:")
                        print(f"         • Anchor row: {row_num}")
                        print(f"         • Anchor col: {col_num}")
                        
                        # Check image data
                        if hasattr(img, '_data'):
                            img_data = img._data()
                            print(f"         • Data size: {len(img_data)} bytes")
                            
                            # Detect format
                            if img_data.startswith(b'\x89PNG'):
                                format_type = 'PNG'
                            elif img_data.startswith(b'\xff\xd8'):
                                format_type = 'JPEG'
                            else:
                                format_type = 'Unknown'
                            print(f"         • Format: {format_type}")
                        
                        # Check if we can get format from image object
                        if hasattr(img, 'format'):
                            print(f"         • Image format: {img.format}")
                        
                    except Exception as e:
                        print(f"         ❌ Error processing image {i+1}: {e}")
            else:
                print(f"   📭 No embedded images found")
        
        print(f"\n📊 Total embedded images found: {total_images}")
        
        if total_images == 0:
            print(f"\n⚠️  NO EMBEDDED IMAGES DETECTED!")
            print(f"   This means the Excel file has images as:")
            print(f"   • External links/URLs (not embedded)")
            print(f"   • Cell values with image URLs")
            print(f"   • Images that are not properly embedded")
            
            print(f"\n🔍 Let's check cell values for image URLs...")
            
            # Check cell values for image URLs
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\n📋 Checking cell values in '{sheet_name}':")
                
                # Check first few rows and columns for image-related data
                for row in range(1, min(6, ws.max_row + 1)):
                    for col in range(1, min(10, ws.max_column + 1)):
                        cell_value = ws.cell(row, col).value
                        if cell_value and isinstance(cell_value, str):
                            if any(keyword in cell_value.lower() for keyword in ['http', 'image', 'photo', 'picture', '.jpg', '.png']):
                                print(f"      Row {row}, Col {col}: {cell_value}")
        
        return total_images > 0
        
    except Exception as e:
        print(f"❌ Error loading workbook: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_service_integration():
    """Test the service integration"""
    print(f"\n🔧 TESTING SERVICE INTEGRATION")
    print("=" * 30)
    
    try:
        # Create test apartment
        client, _ = Client.objects.get_or_create(
            name="OpenpyXL Test Client",
            defaults={'email': 'test@example.com', 'phone': '123-456-7890'}
        )
        
        apartment, created = Apartment.objects.get_or_create(
            name="OpenpyXL Test Apartment",
            defaults={
                'client': client,
                'type': 'furnishing',
                'address': 'Test Address',
                'status': 'Planning',
                'start_date': '2024-01-01',
                'due_date': '2024-12-31',
            }
        )
        
        print(f"✅ Test apartment ready: {apartment.name}")
        
        # Test image extraction service
        service = ProductImportService()
        excel_file = 'sample_products_with_images.xlsx'
        
        if os.path.exists(excel_file):
            print(f"🔍 Testing image extraction service...")
            row_image_map = service._extract_excel_images_with_openpyxl(excel_file, apartment)
            
            print(f"📊 Extraction results:")
            print(f"   • Sheets processed: {len(row_image_map)}")
            
            total_images = 0
            for sheet_name, sheet_images in row_image_map.items():
                print(f"   • {sheet_name}: {len(sheet_images)} images")
                total_images += len(sheet_images)
                
                for row_num, image_path in sheet_images.items():
                    print(f"     - Row {row_num}: {image_path}")
            
            print(f"   • Total images extracted: {total_images}")
            
            return total_images > 0
        else:
            print(f"❌ Excel file not found for service test")
            return False
            
    except Exception as e:
        print(f"❌ Service integration test failed: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 OPENPYXL IMAGE EXTRACTION TEST\n")
    
    try:
        # Test 1: Direct openpyxl extraction
        test1_success = test_openpyxl_extraction()
        
        # Test 2: Service integration
        test2_success = test_service_integration()
        
        print(f"\n📊 TEST RESULTS:")
        print(f"   • OpenpyXL extraction: {'✅ PASS' if test1_success else '❌ FAIL'}")
        print(f"   • Service integration: {'✅ PASS' if test2_success else '❌ FAIL'}")
        
        if test1_success and test2_success:
            print(f"\n🎉 ALL TESTS PASSED!")
            print(f"   Ready to import Excel with embedded images")
        elif not test1_success:
            print(f"\n⚠️  NO EMBEDDED IMAGES FOUND")
            print(f"   The Excel file contains image URLs, not embedded images")
            print(f"   The original URL-based import should work")
        else:
            print(f"\n❌ SOME TESTS FAILED")
            print(f"   Check error messages above")
            
    except Exception as e:
        print(f"\n❌ Test failed: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print(f"\n✅ Test complete!")
